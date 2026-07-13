# -*- coding: utf-8 -*-
"""
需求数据清洗与模板生成系统 - 统一主入口
"""
import sys
import os
import argparse

sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'src'))

from src.pipeline import DataProcessingPipeline
from src.llm_client import llm_client


def read_requirements_from_excel(file_path, logger=None):
    """从Excel文件读取需求数据"""
    requirements = []
    
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
        
        sheets = wb.sheetnames
        if logger:
            logger.info("FileReader", "EXCEL", f"文件包含工作表: {sheets}")
        
        if len(sheets) > 1:
            ws = wb[sheets[1]]
            if logger:
                logger.info("FileReader", "EXCEL", f"使用第2个工作表: {sheets[1]}")
        else:
            ws = wb[sheets[0]]
            if logger:
                logger.info("FileReader", "EXCEL", f"使用唯一工作表: {sheets[0]}")
        
        req_text_col = 2  
        req_id_col = 3    
        
        if logger:
            logger.info("FileReader", "EXCEL", f"使用固定列配置: 需求文本列=C({req_text_col+1}), 序号列=D({req_id_col+1})")
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            req_id = str(row[req_id_col]) if len(row) > req_id_col else ""
            req_text = str(row[req_text_col]) if len(row) > req_text_col else ""
            
            if req_id and req_text and req_id != "None" and req_text != "None":
                requirements.append({
                    "id": req_id.strip(),
                    "text": req_text.strip()
                })
        
        wb.close()
        if logger:
            logger.info("FileReader", "EXCEL", f"从文件读取了 {len(requirements)} 条需求")
        
    except Exception as e:
        if logger:
            logger.error("FileReader", "EXCEL", f"读取Excel文件失败: {str(e)}")
    
    return requirements


def run_demo(pipeline):
    """运行演示模式 - 处理单条需求"""
    print("\n处理示例需求...")
    instance = pipeline.process_single_requirement(
        requirement_text="运行功耗<3mA，休眠功耗<10μA",
        requirement_id="REQ_DEMO_001",
        product_line="MCU",
        chip_info="CS8M320"
    )
    
    if instance:
        print("[OK] 处理成功")
        print("  需求ID: {}".format(instance.requirement_instance_id))
        print("  分类UID: {}".format(instance.category_uid))
        print("  提取变量: {}".format(instance.extracted_variables))


def run_batch(pipeline, data_dir, run_template_matching=True):
    """运行批量处理模式 - 处理data目录中的Excel文件"""
    if not os.path.exists(data_dir):
        print("[ERROR] 数据目录不存在: {}".format(data_dir))
        return
    
    excel_files = []
    for filename in os.listdir(data_dir):
        if filename.endswith('.xlsx'):
            excel_files.append(os.path.join(data_dir, filename))
    
    if not excel_files:
        print("[INFO] 数据目录中没有Excel文件")
        return
    
    logger = pipeline.logger
    
    logger.info("System", "INIT", "需求数据清洗与模板生成系统 - 真实数据处理（生产环境）")
    logger.info("System", "INIT", "系统初始化完成")
    logger.info("System", "INIT", f"使用真实LLM客户端: {llm_client.model_name}")
    
    for excel_file in excel_files:
        filename = os.path.basename(excel_file)
        logger.info("FileProcessor", "BATCH", "====================================================")
        logger.info("FileProcessor", "BATCH", f"处理文件: {filename}")
        logger.info("FileProcessor", "BATCH", "====================================================")
        
        parts = filename.split(" - ")
        product_line = parts[0].strip() if len(parts) > 0 else "UNKNOWN"
        
        if len(parts) > 1:
            chip_info_raw = parts[1].strip().replace(".xlsx", "")
            if " - " in chip_info_raw:
                chip_info_raw = chip_info_raw.split(" - ")[0].strip()
            import re
            chip_info_raw = re.sub(r'\[.*?\]', '', chip_info_raw).strip()
            chip_info = chip_info_raw
        else:
            chip_info = ""
        
        requirements = read_requirements_from_excel(excel_file, logger)
        
        if not requirements:
            logger.warning("FileProcessor", "BATCH", "文件中没有有效需求")
            continue
        
        logger.info("FileProcessor", "BATCH", f"开始批量处理 {len(requirements)} 条需求...")
        instances = pipeline.process_batch(
            requirements=requirements,
            product_line=product_line,
            chip_info=chip_info,
            run_template_matching=run_template_matching
        )
        
        logger.info("FileProcessor", "BATCH", f"处理完成，共 {len(instances)} 条需求")


def main():
    parser = argparse.ArgumentParser(description='需求数据清洗与模板生成系统')
    parser.add_argument('--demo', action='store_true', help='运行演示模式（处理单条示例需求）')
    parser.add_argument('--no-template', action='store_true', help='跳过模板匹配（仅分类）')
    parser.add_argument('--data-dir', type=str, default=None, help='指定数据目录（默认使用 data/ 目录）')

    # 规范JSON驱动模式
    parser.add_argument('--spec-json', type=str, default=None,
                        help='规范JSON路径（自动派生categories和templates）')
    parser.add_argument('--no-spec-provider', action='store_true',
                        help='禁用SpecDataProvider，使用传统JSON文件模式')
    parser.add_argument('--check-spec', action='store_true',
                        help='启动时检查规范文件是否变更')
    parser.add_argument('--auto-sync', action='store_true',
                        help='规范变更时自动同步配置（配合--check-spec使用）')

    args = parser.parse_args()

    print("=" * 60)
    print("需求数据清洗与模板生成系统")
    print("=" * 60)

    base_dir = os.path.dirname(__file__)

    # 确定规范JSON路径
    spec_json_path = args.spec_json
    if not spec_json_path:
        default_spec = os.path.join(base_dir, "data", "config", "芯片需求结构化定义规范V4.0.json")
        if os.path.exists(default_spec):
            spec_json_path = default_spec

    # 规范变更检测
    hash_file = os.path.join(base_dir, "data", "config", ".spec_hash")
    if args.check_spec and spec_json_path and os.path.exists(spec_json_path):
        from src.spec_data_provider import SpecDataProvider
        tmp_provider = SpecDataProvider(
            spec_json_path=spec_json_path,
            templates_json_path=os.path.join(base_dir, "data", "config", "Master_Requirement_Templates.json"),
            existing_category_db_path=os.path.join(base_dir, "data", "config", "categories.dbV4.1.json")
        )
        changed, msg = tmp_provider.check_spec_changed(hash_file)
        if changed:
            print(f"[SPEC] {msg}")
            if args.auto_sync:
                print("[SPEC] 自动同步模式，继续使用最新规范数据...")
                tmp_provider.save_spec_hash(hash_file)
            else:
                print("[SPEC] 提示: 规范文件已变更，建议运行 python script/sync_config.py --sync")
                print("[SPEC] 或使用 --auto-sync 自动同步")
        else:
            print(f"[SPEC] {msg}")

        stats = tmp_provider.get_stats()
        print(f"[SPEC] Categories: {stats['categories_l1']}L1/{stats['categories_l2']}L2/{stats['categories_l3']}L3")
        print(f"[SPEC] Templates: {stats['total_templates']} (existing: {stats.get('existing_templates', 0)}, auto: {stats.get('auto_generated_templates', 0)})")

    # 初始化Pipeline
    use_spec = not args.no_spec_provider and spec_json_path and os.path.exists(spec_json_path)

    if use_spec:
        print(f"\n[INFO] 使用规范JSON驱动模式: {spec_json_path}")
        pipeline = DataProcessingPipeline(
            category_db_path=os.path.join(base_dir, "data", "config", "categories.dbV4.1.json"),
            templates_path=os.path.join(base_dir, "data", "config", "Master_Requirement_Templates.json"),
            llm_client=llm_client,
            audit_records_dir="Audit_Records",
            output_dir="output",
            template_library_dir="output/library",
            spec_json_path=spec_json_path,
            use_spec_provider=True
        )
        # 保存规范哈希
        if args.auto_sync:
            pipeline.spec_provider.save_spec_hash(hash_file)
    else:
        print("\n[INFO] 使用传统JSON文件模式")
        pipeline = DataProcessingPipeline(
            category_db_path=os.path.join(base_dir, "data", "config", "categories.dbV4.1.json"),
            templates_path=os.path.join(base_dir, "data", "config", "Master_Requirement_Templates.json"),
            llm_client=llm_client,
            audit_records_dir="Audit_Records",
            output_dir="output",
            template_library_dir="output/library"
        )

    print("[OK] 系统初始化完成")
    print("[INFO] 使用真实LLM客户端: {}".format(llm_client.model_name))

    if args.demo:
        run_demo(pipeline)
    else:
        data_dir = args.data_dir if args.data_dir else os.path.join(base_dir, 'data')
        run_batch(pipeline, data_dir, run_template_matching=not args.no_template)

    print("\n" + "=" * 60)
    print("处理完成")
    print("=" * 60)


if __name__ == "__main__":
    main()