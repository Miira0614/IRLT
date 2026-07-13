# -*- coding: utf-8 -*-
"""可视化模块 - 负责生成Excel预览和追踪链关系图"""

from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment
)
from openpyxl.utils import get_column_letter
from .data_models import RequirementInstance


class VisualizationModule:
    """可视化模块"""

    # 颜色定义
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    AI_GENERATED_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    MANUAL_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    PENDING_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    L1_INDENT = 0
    L2_INDENT = 1
    L3_INDENT = 2

    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def __init__(
        self,
        output_dir: str = "output",
        category_db=None,
        logger=None
    ):
        """
        初始化可视化模块

        Args:
            output_dir: 输出目录
            category_db: 分类数据库实例（用于解析分类UID为名称）
            logger: 日志记录器
        """
        self.output_dir = output_dir
        self.category_db = category_db
        self.logger = logger

    def render_to_excel(
        self,
        instances: List[RequirementInstance],
        output_filename: str,
        title: str = "需求数据预览"
    ) -> str:
        """
        渲染需求数据到Excel文件

        Args:
            instances: 需求实例列表
            output_filename: 输出文件名
            title: 报表标题

        Returns:
            生成的文件路径
        """
        wb = Workbook()
        
        # 创建第一个工作表：需求详情
        ws1 = wb.active
        ws1.title = "需求详情"
        self._write_requirement_details(ws1, instances)
        
        # 创建第二个工作表：统计概览
        ws2 = wb.create_sheet(title="统计概览")
        self._write_statistics_overview(ws2, instances)
        
        # 保存文件
        output_path = f"{self.output_dir}/{output_filename}"
        wb.save(output_path)
        
        if self.logger:
            self.logger.info(
                "Visualization",
                "ExcelRenderer",
                f"成功生成预览Excel: {output_path}, 包含{len(instances)}条需求"
            )
        
        return output_path

    def _write_requirement_details(
        self,
        ws,
        instances: List[RequirementInstance]
    ):
        """写入需求详情表"""
        # 设置列宽
        column_widths = [12, 15, 50, 20, 20, 12, 15, 15, 12, 12, 30]
        for i, width in enumerate(column_widths):
            ws.column_dimensions[get_column_letter(i+1)].width = width
        
        # 写入表头
        headers = [
            "需求ID", "层级", "需求文本", "分类名称", "分类UID", "置信度",
            "父级需求", "子级需求", "生成类型", "审核状态", "产品线"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 写入数据
        row = 2
        for inst in instances:
            # 获取分类名称
            cat_name = ""
            if self.category_db and inst.category_uid:
                cat = self.category_db.get_category(inst.category_uid)
                if cat:
                    cat_name = cat.name
            
            # 获取父级和子级需求
            parent_id = inst.instance_trace_chain.get("parent_requirement_id", "")
            child_ids = inst.instance_trace_chain.get("child_requirement_ids", [])
            child_ids_display = ", ".join(child_ids) if child_ids else ""
            
            # 获取生成类型显示
            gen_type = "AI生成" if inst.generation_type == "ai_generated" else "文件导入"
            
            # 获取审核状态显示
            review_status = "待审核" if inst.review_status == "pending_review" else "已审核"
            
            # 写入行数据
            row_data = [
                inst.requirement_instance_id,
                inst.requirement_type,
                inst.requirement_text,
                cat_name,
                inst.category_uid,
                inst.confidence,  # 置信度
                parent_id,
                child_ids_display,
                gen_type,
                review_status,
                inst.product_line
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                if inst.generation_type == "ai_generated":
                    cell.fill = self.AI_GENERATED_FILL
                elif inst.review_status == "pending_review":
                    cell.fill = self.PENDING_FILL
            
            row += 1

    def _write_statistics_overview(
        self,
        ws,
        instances: List[RequirementInstance]
    ):
        """写入统计概览表 - 按L3为核心组织，显示完整追溯链"""
        # 设置列宽
        column_widths = [35, 30, 30, 20, 45, 20, 45, 20, 45, 50]
        for i, width in enumerate(column_widths):
            ws.column_dimensions[get_column_letter(i+1)].width = width
        
        # 标题
        ws.cell(row=1, column=1, value="需求统计概览").font = Font(bold=True, size=14)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
        
        # 表头（新格式）
        headers = [
            "需求维度", "能力域", "需求项", 
            "需求ID", "L1客户需求", 
            "需求ID", "L2初始需求", 
            "需求ID", "L3系统需求", 
            "需求追溯链"
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # ✨【强关联模式】按L3为核心组织数据，L1和L2对L3强关联
        # 构建ID到实例的映射
        id_to_instance = {inst.requirement_instance_id: inst for inst in instances}
        
        # 按三级分类分组
        category_map = {}
        for inst in instances:
            cat_uid = inst.category_uid
            if cat_uid and cat_uid != "cat_failed":
                cat = self.category_db.get_category(cat_uid) if self.category_db else None
                if cat:
                    # 获取完整的分类层级路径
                    l1_id, l1_name = "", ""
                    l2_id, l2_name = "", ""
                    l3_id, l3_name = cat.id, cat.name
                    
                    # 向上查找父级
                    if cat.parent_uid:
                        parent = self.category_db.get_category(cat.parent_uid)
                        if parent:
                            l2_id, l2_name = parent.id, parent.name
                            if parent.parent_uid:
                                grandparent = self.category_db.get_category(parent.parent_uid)
                                if grandparent:
                                    l1_id, l1_name = grandparent.id, grandparent.name
                    
                    key = (l1_id, l2_id, l3_id)
                    if key not in category_map:
                        category_map[key] = {
                            "l1": f"{l1_id} - {l1_name}" if l1_id else "",
                            "l2": f"{l2_id} - {l2_name}" if l2_id else "",
                            "l3": f"{l3_id} - {l3_name}" if l3_id else "",
                            "L1": [],
                            "L2": [],
                            "L3": []
                        }
                    category_map[key][inst.requirement_type].append(inst)
    
        # 写入数据（以L3为核心，L3在哪一行，对应的L1/L2就放在哪一行）
        row = 4
        total_rows = 0
        
        # 记录上一行的内容，用于判断是否重复
        prev_l1_id = ""
        prev_l2_id = ""
        
        for key in sorted(category_map.keys()):
            data = category_map[key]
            
            # 获取所有L3需求
            l3_instances = data["L3"]
            
            # 记录已关联的L1/L2实例ID（用于后续查找孤立需求）
            associated_l1_ids = set()
            associated_l2_ids = set()
            
            if l3_instances:
                # 有L3需求，按L3数量决定行数
                for l3_inst in l3_instances:
                    # 获取L3对应的父级L2
                    l2_inst = None
                    # 获取L3对应的父级L2
                    l2_id = ""
                    l2_text = ""
                    parent_id = l3_inst.instance_trace_chain.get("parent_requirement_id")
                    if parent_id and parent_id in id_to_instance:
                        l2_inst = id_to_instance[parent_id]
                        if l2_inst.requirement_type == "L2":
                            l2_id = l2_inst.requirement_instance_id
                            l2_text = l2_inst.requirement_text
                            associated_l2_ids.add(l2_id)
                    
                    # 获取L2对应的父级L1
                    l1_inst = None
                    l1_id = ""
                    # 适配新的instance_trace_chain结构，更新L1父级需求获取逻辑
                    l1_text = ""
                    if l2_inst:
                        l1_parent_id = l2_inst.instance_trace_chain.get("parent_requirement_id")
                        if l1_parent_id and l1_parent_id in id_to_instance:
                            l1_inst = id_to_instance[l1_parent_id]
                            if l1_inst.requirement_type == "L1":
                                l1_id = l1_inst.requirement_instance_id
                                l1_text = l1_inst.requirement_text
                                associated_l1_ids.add(l1_id)
                    
                    # 构建需求追溯链
                    trace_chain_parts = []
                    if l1_id:
                        trace_chain_parts.append(l1_id)
                    if l2_id:
                        trace_chain_parts.append(l2_id)
                    if l3_inst.requirement_instance_id:
                        trace_chain_parts.append(l3_inst.requirement_instance_id)
                    trace_chain = "--".join(trace_chain_parts)
                    
                    # 写入分类信息（重复显示，确保每一行都完整）
                    ws.cell(row=row, column=1, value=data["l1"])
                    ws.cell(row=row, column=2, value=data["l2"])
                    ws.cell(row=row, column=3, value=data["l3"])
                    
                    # ✨ 优化：重复的L1内容只显示一次
                    if l1_id != prev_l1_id:
                        # 写入L1需求ID
                        ws.cell(row=row, column=4, value=l1_id)
                        
                        # 写入L1需求文本（带颜色渲染）
                        l1_cell = ws.cell(row=row, column=5, value=l1_text)
                        if l1_inst:
                            if l1_inst.generation_type == "ai_generated":
                                l1_cell.fill = self.AI_GENERATED_FILL
                            elif l1_inst.review_status == "pending_review":
                                l1_cell.fill = self.PENDING_FILL
                        l1_cell.alignment = Alignment(wrap_text=True, vertical="top")
                        prev_l1_id = l1_id
                    else:
                        # 重复内容，留空
                        ws.cell(row=row, column=4, value="")
                        ws.cell(row=row, column=5, value="")
                    
                    # ✨ 优化：重复的L2内容只显示一次
                    if l2_id != prev_l2_id:
                        # 写入L2需求ID
                        ws.cell(row=row, column=6, value=l2_id)
                        
                        # 写入L2需求文本（带颜色渲染）
                        l2_cell = ws.cell(row=row, column=7, value=l2_text)
                        if l2_inst:
                            if l2_inst.generation_type == "ai_generated":
                                l2_cell.fill = self.AI_GENERATED_FILL
                            elif l2_inst.review_status == "pending_review":
                                l2_cell.fill = self.PENDING_FILL
                        l2_cell.alignment = Alignment(wrap_text=True, vertical="top")
                        prev_l2_id = l2_id
                    else:
                        # 重复内容，留空
                        ws.cell(row=row, column=6, value="")
                        ws.cell(row=row, column=7, value="")
                    
                    # 写入L3需求ID
                    ws.cell(row=row, column=8, value=l3_inst.requirement_instance_id)
                    
                    # 写入L3需求文本（带颜色渲染）
                    l3_cell = ws.cell(row=row, column=9, value=l3_inst.requirement_text)
                    if l3_inst.generation_type == "ai_generated":
                        l3_cell.fill = self.AI_GENERATED_FILL
                    elif l3_inst.review_status == "pending_review":
                        l3_cell.fill = self.PENDING_FILL
                    l3_cell.alignment = Alignment(wrap_text=True, vertical="top")
                    
                    # 写入需求追溯链
                    ws.cell(row=row, column=10, value=trace_chain)
                    
                    row += 1
                    total_rows += 1
                
                # ✨【修复】额外渲染孤立的L1/L2需求（没有被任何L3关联的）
                # 找出孤立的L1需求
                isolated_l1 = [l1 for l1 in data["L1"] if l1.requirement_instance_id not in associated_l1_ids]
                # 找出孤立的L2需求（包括没有L1父级的L2）
                isolated_l2 = [l2 for l2 in data["L2"] if l2.requirement_instance_id not in associated_l2_ids]
                
                # 渲染孤立的L1需求
                for l1_inst in isolated_l1:
                    # 写入分类信息
                    ws.cell(row=row, column=1, value=data["l1"])
                    ws.cell(row=row, column=2, value=data["l2"])
                    ws.cell(row=row, column=3, value=data["l3"])
                    
                    # 写入L1
                    ws.cell(row=row, column=4, value=l1_inst.requirement_instance_id)
                    l1_cell = ws.cell(row=row, column=5, value=l1_inst.requirement_text)
                    if l1_inst.generation_type == "ai_generated":
                        l1_cell.fill = self.AI_GENERATED_FILL
                    elif l1_inst.review_status == "pending_review":
                        l1_cell.fill = self.PENDING_FILL
                    l1_cell.alignment = Alignment(wrap_text=True, vertical="top")
                    
                    # 写入追溯链
                    ws.cell(row=row, column=10, value=l1_inst.requirement_instance_id)
                    
                    row += 1
                    total_rows += 1
                
                # 渲染孤立的L2需求
                for l2_inst in isolated_l2:
                    # 获取L2对应的父级L1（如果存在）
                    l1_id = ""
                    l1_text = ""
                    l1_parent_id = l2_inst.instance_trace_chain.get("parent_requirement_id")
                    if l1_parent_id and l1_parent_id in id_to_instance:
                        l1_inst = id_to_instance[l1_parent_id]
                        if l1_inst.requirement_type == "L1":
                            l1_id = l1_inst.requirement_instance_id
                            l1_text = l1_inst.requirement_text
                    
                    # 写入分类信息
                    ws.cell(row=row, column=1, value=data["l1"])
                    ws.cell(row=row, column=2, value=data["l2"])
                    ws.cell(row=row, column=3, value=data["l3"])
                    
                    # 写入L1（如果有）
                    ws.cell(row=row, column=4, value=l1_id)
                    l1_cell = ws.cell(row=row, column=5, value=l1_text)
                    l1_cell.alignment = Alignment(wrap_text=True, vertical="top")
                    
                    # 写入L2
                    ws.cell(row=row, column=6, value=l2_inst.requirement_instance_id)
                    l2_cell = ws.cell(row=row, column=7, value=l2_inst.requirement_text)
                    if l2_inst.generation_type == "ai_generated":
                        l2_cell.fill = self.AI_GENERATED_FILL
                    elif l2_inst.review_status == "pending_review":
                        l2_cell.fill = self.PENDING_FILL
                    l2_cell.alignment = Alignment(wrap_text=True, vertical="top")
                    
                    # 写入追溯链
                    trace_chain_parts = []
                    if l1_id:
                        trace_chain_parts.append(l1_id)
                    trace_chain_parts.append(l2_inst.requirement_instance_id)
                    ws.cell(row=row, column=10, value="--".join(trace_chain_parts))
                    
                    row += 1
                    total_rows += 1
            
            else:
                # 没有L3需求，按最大数量决定行数（兼容旧逻辑）
                max_count = max(len(data["L1"]), len(data["L2"]))
                for i in range(max_count):
                    # 写入分类信息
                    ws.cell(row=row+i, column=1, value=data["l1"])
                    ws.cell(row=row+i, column=2, value=data["l2"])
                    ws.cell(row=row+i, column=3, value=data["l3"])
                    
                    # 写入L1
                    if i < len(data["L1"]):
                        l1_inst = data["L1"][i]
                        ws.cell(row=row+i, column=4, value=l1_inst.requirement_instance_id)
                        l1_cell = ws.cell(row=row+i, column=5, value=l1_inst.requirement_text)
                        if l1_inst.generation_type == "ai_generated":
                            l1_cell.fill = self.AI_GENERATED_FILL
                        elif l1_inst.review_status == "pending_review":
                            l1_cell.fill = self.PENDING_FILL
                        l1_cell.alignment = Alignment(wrap_text=True, vertical="top")
                    
                    # 写入L2
                    if i < len(data["L2"]):
                        l2_inst = data["L2"][i]
                        ws.cell(row=row+i, column=6, value=l2_inst.requirement_instance_id)
                        l2_cell = ws.cell(row=row+i, column=7, value=l2_inst.requirement_text)
                        if l2_inst.generation_type == "ai_generated":
                            l2_cell.fill = self.AI_GENERATED_FILL
                        elif l2_inst.review_status == "pending_review":
                            l2_cell.fill = self.PENDING_FILL
                        l2_cell.alignment = Alignment(wrap_text=True, vertical="top")
                
                row += max_count
        
        # 添加图例说明
        row += 2
        ws.cell(row=row, column=1, value="图例：").font = Font(bold=True)
        legend_items = [
            ("AI生成", self.AI_GENERATED_FILL),
            ("待审核", self.PENDING_FILL)
        ]
        
        for i, (label, fill) in enumerate(legend_items):
            cell = ws.cell(row=row, column=2+i*2, value=label)
            cell.fill = fill

    def render_trace_chain_diagram(
        self,
        instances: List[RequirementInstance],
        output_filename: str
    ) -> str:
        """
        生成追踪链关系图（文本格式）

        Args:
            instances: 需求实例列表
            output_filename: 输出文件名

        Returns:
            生成的文件路径
        """
        output_path = f"{self.output_dir}/{output_filename}"
        
        # 按层级分组
        l1_instances = [inst for inst in instances if inst.requirement_type == "L1"]
        l2_instances = [inst for inst in instances if inst.requirement_type == "L2"]
        l3_instances = [inst for inst in instances if inst.requirement_type == "L3"]
        
        # 创建ID到实例的映射
        id_to_instance = {inst.requirement_instance_id: inst for inst in instances}
        
        # 去除重复的实例（根据需求ID去重）
        seen_ids = set()
        unique_l1 = []
        for inst in l1_instances:
            if inst.requirement_instance_id not in seen_ids:
                seen_ids.add(inst.requirement_instance_id)
                unique_l1.append(inst)
        
        seen_ids = set()
        unique_l2 = []
        for inst in l2_instances:
            if inst.requirement_instance_id not in seen_ids:
                seen_ids.add(inst.requirement_instance_id)
                unique_l2.append(inst)
        
        seen_ids = set()
        unique_l3 = []
        for inst in l3_instances:
            if inst.requirement_instance_id not in seen_ids:
                seen_ids.add(inst.requirement_instance_id)
                unique_l3.append(inst)
        
        with open(output_path, "w", encoding="utf-8") as f:
            f.write("=" * 80 + "\n")
            f.write("需求追踪链关系图\n")
            f.write("=" * 80 + "\n\n")
            
            # 按层级输出
            f.write("【L1 - 客户需求】\n")
            f.write("-" * 80 + "\n")
            for inst in unique_l1:
                f.write(f"  {inst.requirement_instance_id}\n")
                f.write(f"    文本: {inst.requirement_text}\n")
                f.write(f"    子级: {inst.instance_trace_chain.get('child_requirement_ids', [])}\n\n")
            
            f.write("\n【L2 - 初始需求】\n")
            f.write("-" * 80 + "\n")
            for inst in unique_l2:
                f.write(f"  {inst.requirement_instance_id}\n")
                f.write(f"    文本: {inst.requirement_text}\n")
                f.write(f"    父级: {inst.instance_trace_chain.get('parent_requirement_id', '')}\n")
                f.write(f"    子级: {inst.instance_trace_chain.get('child_requirement_ids', [])}\n\n")
            
            f.write("\n【L3 - 系统需求】\n")
            f.write("-" * 80 + "\n")
            for inst in unique_l3:
                f.write(f"  {inst.requirement_instance_id}\n")
                f.write(f"    文本: {inst.requirement_text}\n")
                f.write(f"    父级: {inst.instance_trace_chain.get('parent_requirement_id', '')}\n\n")
        
        if self.logger:
            self.logger.info(
                "Visualization",
                "TraceChainDiagram",
                f"成功生成追踪链关系图: {output_path}"
            )
        
        return output_path

    def generate_summary_stats(
        self,
        instances: List[RequirementInstance]
    ) -> Dict[str, Any]:
        """
        生成汇总统计信息

        Args:
            instances: 需求实例列表

        Returns:
            统计信息字典
        """
        stats = {
            "L1_count": sum(1 for inst in instances if inst.requirement_type == "L1"),
            "L2_count": sum(1 for inst in instances if inst.requirement_type == "L2"),
            "L3_count": sum(1 for inst in instances if inst.requirement_type == "L3"),
            "ai_generated_count": sum(1 for inst in instances if inst.generation_type == "ai_generated"),
            "manual_count": sum(1 for inst in instances if inst.generation_type != "ai_generated"),
            "pending_review_count": sum(1 for inst in instances if inst.review_status == "pending_review"),
            "approved_count": sum(1 for inst in instances if inst.review_status == "approved")
        }
        
        return stats
