import sys
import os

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
sys.path.insert(0, PROJECT_ROOT)
sys.path.insert(0, os.path.join(PROJECT_ROOT, 'src'))


def _init_pipeline(data_dir=None, use_spec_provider=True):
    """初始化数据处理流水线"""
    from src.pipeline import DataProcessingPipeline
    from src.llm_client import llm_client

    base_dir = PROJECT_ROOT

    spec_json_path = os.path.join(base_dir, "data", "config", "芯片需求结构化定义规范V4.1.json")
    if not os.path.exists(spec_json_path):
        spec_json_path = os.path.join(base_dir, "data", "config", "芯片需求结构化定义规范V4.0.json")

    pipeline = DataProcessingPipeline(
        category_db_path=os.path.join(base_dir, "data", "config", "categories.dbV4.1.json"),
        templates_path=os.path.join(base_dir, "data", "config", "Master_Requirement_Templates.json"),
        llm_client=llm_client,
        audit_records_dir="Audit_Records",
        output_dir="output",
        template_library_dir="output/library",
        spec_json_path=spec_json_path if os.path.exists(spec_json_path) else None,
        use_spec_provider=use_spec_provider and os.path.exists(spec_json_path)
    )
    return pipeline


def process_single_requirement(requirement_text, requirement_id="REQ_AUTO",
                               product_line="MCU", chip_info=""):
    """
    处理单条需求
    
    Args:
        requirement_text: 需求文本
        requirement_id: 需求ID（可选）
        product_line: 产品线（可选，默认MCU）
        chip_info: 芯片信息（可选）
    
    Returns:
        RequirementInstance 对象，包含分类结果、模板匹配结果等
    """
    pipeline = _init_pipeline()
    
    instance = pipeline.process_single_requirement(
        requirement_text=requirement_text,
        requirement_id=requirement_id,
        product_line=product_line,
        chip_info=chip_info
    )
    
    return instance


def process_batch(data_dir=None, run_template_matching=True):
    """
    批量处理Excel文件中的需求
    
    Args:
        data_dir: 数据目录（可选，默认 data/）
        run_template_matching: 是否运行模板匹配（可选，默认True）
    
    Returns:
        处理后的需求实例列表
    """
    pipeline = _init_pipeline()
    
    if data_dir is None:
        data_dir = os.path.join(PROJECT_ROOT, 'data')
    
    excel_files = []
    for filename in os.listdir(data_dir):
        if filename.endswith('.xlsx'):
            excel_files.append(os.path.join(data_dir, filename))
    
    if not excel_files:
        print("[INFO] 数据目录中没有Excel文件")
        return []
    
    all_instances = []
    for excel_file in excel_files:
        filename = os.path.basename(excel_file)
        print(f"处理文件: {filename}")
        
        parts = filename.split(" - ")
        product_line = parts[0].strip() if len(parts) > 0 else "UNKNOWN"
        
        if len(parts) > 1:
            chip_info_raw = parts[1].strip().replace(".xlsx", "")
            import re
            chip_info_raw = re.sub(r'\[.*?\]', '', chip_info_raw).strip()
            chip_info = chip_info_raw
        else:
            chip_info = ""
        
        requirements = _read_requirements_from_excel(excel_file)
        
        if not requirements:
            print(f"文件 {filename} 中没有有效需求")
            continue
        
        instances = pipeline.process_batch(
            requirements=requirements,
            product_line=product_line,
            chip_info=chip_info,
            run_template_matching=run_template_matching
        )
        all_instances.extend(instances)
    
    return all_instances


def _read_requirements_from_excel(file_path):
    """从Excel文件读取需求数据"""
    requirements = []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
        sheets = wb.sheetnames
        ws = wb[sheets[0]]
        
        req_text_col = 2
        req_id_col = 3
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            req_id = str(row[req_id_col]) if len(row) > req_id_col else ""
            req_text = str(row[req_text_col]) if len(row) > req_text_col else ""
            
            if req_id and req_text and req_id != "None" and req_text != "None":
                requirements.append({
                    "id": req_id.strip(),
                    "text": req_text.strip()
                })
        wb.close()
    except Exception as e:
        print(f"读取Excel文件失败: {str(e)}")
    
    return requirements


def generate_trace_matrix(input_dir=None, output_dir=None):
    """
    生成需求追溯链矩阵（聚合链矩阵）
    
    Args:
        input_dir: 输入目录（原始需求Excel，可选，默认 raw_data/）
        output_dir: 输出目录（可选，默认 raw_out/）
    
    Returns:
        生成的文件路径
    """
    if input_dir is None:
        input_dir = os.path.join(PROJECT_ROOT, 'raw_data')
    
    if output_dir is None:
        output_dir = os.path.join(PROJECT_ROOT, 'raw_out')
    
    os.makedirs(output_dir, exist_ok=True)
    
    sys.path.insert(0, os.path.join(PROJECT_ROOT, 'script'))
    from script.需求表格预处理 import main_orchestrator
    
    import script.需求表格预处理 as preprocessor
    preprocessor.INPUT_FOLDER = input_dir
    preprocessor.OUTPUT_FOLDER = output_dir
    
    main_orchestrator()
    
    output_files = [f for f in os.listdir(output_dir) if f.endswith('.xlsx')]
    if output_files:
        return os.path.join(output_dir, output_files[0])
    return None


def get_stats():
    """
    获取系统统计信息
    
    Returns:
        包含分类数、模板数等统计信息的字典
    """
    import sys
    sys.path.insert(0, PROJECT_ROOT)
    sys.path.insert(0, os.path.join(PROJECT_ROOT, 'src'))
    
    from src.spec_data_provider import SpecDataProvider
    
    base_dir = PROJECT_ROOT
    spec_json_path = os.path.join(base_dir, "data", "config", "芯片需求结构化定义规范V4.1.json")
    
    if not os.path.exists(spec_json_path):
        spec_json_path = os.path.join(base_dir, "data", "config", "芯片需求结构化定义规范V4.0.json")
    
    if not os.path.exists(spec_json_path):
        return {"error": "规范文件不存在"}
    
    provider = SpecDataProvider(
        spec_json_path=spec_json_path,
        templates_json_path=os.path.join(base_dir, "data", "config", "Master_Requirement_Templates.json"),
        existing_category_db_path=os.path.join(base_dir, "data", "config", "categories.dbV4.1.json")
    )
    
    return provider.get_stats()


def run_demo():
    """运行演示模式"""
    print("=" * 60)
    print("需求追溯链 Agent - 演示模式")
    print("=" * 60)
    
    stats = get_stats()
    if "error" not in stats:
        print(f"分类数: {stats['categories_l1']} L1 / {stats['categories_l2']} L2 / {stats['categories_l3']} L3")
        print(f"模板数: {stats['total_templates']}")
    
    print("\n处理示例需求...")
    instance = process_single_requirement(
        requirement_text="运行功耗<3mA，休眠功耗<10μA",
        requirement_id="REQ_DEMO_001",
        product_line="MCU",
        chip_info="CS8M320"
    )
    
    if instance:
        print(f"[OK] 处理成功")
        print(f"  需求ID: {instance.requirement_instance_id}")
        print(f"  分类UID: {instance.category_uid}")
        print(f"  需求层级: {instance.requirement_type}")
        print(f"  提取变量: {instance.extracted_variables}")
        print(f"  匹配模板: {instance.matched_template_id}")
    
    print("\n" + "=" * 60)
    print("演示完成")
    print("=" * 60)


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description='需求追溯链 Agent')
    parser.add_argument('--demo', action='store_true', help='运行演示模式')
    parser.add_argument('--stats', action='store_true', help='显示系统统计信息')
    parser.add_argument('--process', action='store_true', help='批量处理数据')
    parser.add_argument('--trace', action='store_true', help='生成追溯链矩阵')
    parser.add_argument('--single', type=str, help='处理单条需求')
    
    args = parser.parse_args()
    
    if args.demo:
        run_demo()
    elif args.stats:
        stats = get_stats()
        print(stats)
    elif args.process:
        instances = process_batch()
        print(f"处理完成，共 {len(instances)} 条需求")
    elif args.trace:
        result = generate_trace_matrix()
        print(f"追溯链矩阵已生成: {result}")
    elif args.single:
        instance = process_single_requirement(args.single)
        print(f"分类UID: {instance.category_uid}")
        print(f"需求层级: {instance.requirement_type}")
    else:
        parser.print_help()
