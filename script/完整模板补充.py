import json
import os
import openpyxl
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter
import time

def parse_version(version_str):
    if not version_str:
        return ()
    parts = version_str.split('.')
    try:
        return tuple(int(p) for p in parts)
    except:  # noqa: E722
        return ()

def extract_identifier(text):
    if not text:
        return None
    if isinstance(text, str):
        parts = text.split('-')[0].strip()
        if '.' in parts:
            try:
                parse_version(parts)
                return parts
            except:  # noqa: E722
                return None
    return None

def has_L_prefix(text):
    if not text:
        return False
    if isinstance(text, str):
        parts = text.split('-')[0].strip()
        return parts.startswith('L')

def get_fill_info(cell):
    fill = cell.fill
    if fill and fill.start_color and fill.start_color.rgb:
        rgb = fill.start_color.rgb
        return rgb
    return None

def process_excel_file(excel_path, spec_map, out_dir):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.worksheets[1]
    
    data_start_row = 4
    
    original_rows_by_id = {}
    all_original_rows = []
    original_only_ids = []
    
    for row in range(data_start_row, ws.max_row + 1):
        c_val = ws[f'C{row}'].value
        identifier = extract_identifier(c_val)
        
        d_val = ws[f'D{row}'].value
        f_val = ws[f'F{row}'].value
        h_val = ws[f'H{row}'].value
        
        row_data = {
            'row': row,
            'identifier': identifier,
            'A': ws[f'A{row}'].value,
            'B': ws[f'B{row}'].value,
            'C': ws[f'C{row}'].value,
            'D': d_val,
            'E': ws[f'E{row}'].value,
            'F': f_val,
            'G': ws[f'G{row}'].value,
            'H': h_val,
            'I': ws[f'I{row}'].value,
            'J': ws[f'J{row}'].value,
            'E_fill': get_fill_info(ws[f'E{row}']),
            'G_fill': get_fill_info(ws[f'G{row}']),
            'I_fill': get_fill_info(ws[f'I{row}']),
            'D_has_L': has_L_prefix(d_val),
            'F_has_L': has_L_prefix(f_val),
            'H_has_L': has_L_prefix(h_val),
        }
        
        all_original_rows.append(row_data)
        
        if identifier:
            if identifier not in original_rows_by_id:
                original_rows_by_id[identifier] = []
                if identifier not in spec_map:
                    original_only_ids.append(identifier)
            original_rows_by_id[identifier].append(row_data)
    
    all_spec_ids = sorted(spec_map.keys(), key=parse_version)
    original_only_ids = sorted(original_only_ids, key=parse_version)
    
    all_ids = []
    spec_idx = 0
    orig_idx = 0
    
    while spec_idx < len(all_spec_ids) or orig_idx < len(original_only_ids):
        spec_id = all_spec_ids[spec_idx] if spec_idx < len(all_spec_ids) else None
        orig_id = original_only_ids[orig_idx] if orig_idx < len(original_only_ids) else None
        
        if spec_id is None:
            all_ids.append(orig_id)
            orig_idx += 1
        elif orig_id is None:
            all_ids.append(spec_id)
            spec_idx += 1
        else:
            spec_tuple = parse_version(spec_id)
            orig_tuple = parse_version(orig_id)
            
            if spec_tuple < orig_tuple:
                all_ids.append(spec_id)
                spec_idx += 1
            else:
                all_ids.append(orig_id)
                orig_idx += 1
    
    new_rows_data = []
    
    for current_id in all_ids:
        if current_id in spec_map:
            spec_item = spec_map[current_id]
            
            if current_id in original_rows_by_id:
                for orig_row in original_rows_by_id[current_id]:
                    new_rows_data.append({
                        'identifier': current_id,
                        'A': f"{spec_item['维度序号']} - {spec_item['维度']}",
                        'B': f"{spec_item['能力域序号']} - {spec_item['能力域']}",
                        'C': f"{spec_item['需求项序号']} - {spec_item['需求项']}",
                        'D': orig_row['D'],
                        'E': orig_row['E'],
                        'F': orig_row['F'],
                        'G': orig_row['G'],
                        'H': orig_row['H'],
                        'I': orig_row['I'],
                        'J': orig_row['J'],
                        'E_fill': orig_row['E_fill'],
                        'G_fill': orig_row['G_fill'],
                        'I_fill': orig_row['I_fill'],
                        'D_has_L': orig_row['D_has_L'],
                        'F_has_L': orig_row['F_has_L'],
                        'H_has_L': orig_row['H_has_L'],
                        'is_new': False,
                    })
            else:
                new_rows_data.append({
                    'identifier': current_id,
                    'A': f"{spec_item['维度序号']} - {spec_item['维度']}",
                    'B': f"{spec_item['能力域序号']} - {spec_item['能力域']}",
                    'C': f"{spec_item['需求项序号']} - {spec_item['需求项']}",
                    'D': f"L1_{current_id}",
                    'E': spec_item.get('原始需求', ''),
                    'F': f"L2_{current_id}",
                    'G': spec_item.get('初始需求', ''),
                    'H': f"L3_{current_id}",
                    'I': spec_item.get('系统需求', ''),
                    'J': '',
                    'E_fill': None,
                    'G_fill': None,
                    'I_fill': None,
                    'D_has_L': True,
                    'F_has_L': True,
                    'H_has_L': True,
                    'is_new': True,
                })
        else:
            for orig_row in original_rows_by_id[current_id]:
                new_rows_data.append({
                    'identifier': current_id,
                    'A': orig_row['A'],
                    'B': orig_row['B'],
                    'C': orig_row['C'],
                    'D': orig_row['D'],
                    'E': orig_row['E'],
                    'F': orig_row['F'],
                    'G': orig_row['G'],
                    'H': orig_row['H'],
                    'I': orig_row['I'],
                    'J': orig_row['J'],
                    'E_fill': orig_row['E_fill'],
                    'G_fill': orig_row['G_fill'],
                    'I_fill': orig_row['I_fill'],
                    'D_has_L': orig_row['D_has_L'],
                    'F_has_L': orig_row['F_has_L'],
                    'H_has_L': orig_row['H_has_L'],
                    'is_new': False,
                })
    
    light_green_fill = PatternFill(start_color='C4D79B', end_color='C4D79B', fill_type='solid')
    ai_generated_fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
    pending_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    wrap_alignment = Alignment(wrap_text=True)
    
    max_col = ws.max_column
    for row in range(data_start_row, ws.max_row + 1):
        for col in range(1, max_col + 1):
            col_letter = get_column_letter(col)
            ws[f'{col_letter}{row}'].value = None
            cell = ws[f'{col_letter}{row}']
            cell.fill = openpyxl.styles.PatternFill(patternType=None, start_color=None, end_color=None)
    
    for idx, row_data in enumerate(new_rows_data):
        row_num = data_start_row + idx
        
        ws[f'A{row_num}'].value = row_data['A']
        ws[f'B{row_num}'].value = row_data['B']
        ws[f'C{row_num}'].value = row_data['C']
        ws[f'D{row_num}'].value = row_data['D']
        ws[f'E{row_num}'].value = row_data['E']
        ws[f'F{row_num}'].value = row_data['F']
        ws[f'G{row_num}'].value = row_data['G']
        ws[f'H{row_num}'].value = row_data['H']
        ws[f'I{row_num}'].value = row_data['I']
        ws[f'J{row_num}'].value = row_data['J']
        
        for col in range(1, max_col + 1):
            col_letter = get_column_letter(col)
            ws[f'{col_letter}{row_num}'].alignment = wrap_alignment
        
        if row_data['is_new']:
            if isinstance(row_data['D'], str) and row_data['D'].strip() and row_data['E']:
                ws[f'E{row_num}'].fill = light_green_fill
            if isinstance(row_data['F'], str) and row_data['F'].strip() and row_data['G']:
                ws[f'G{row_num}'].fill = light_green_fill
            if isinstance(row_data['H'], str) and row_data['H'].strip() and row_data['I']:
                ws[f'I{row_num}'].fill = light_green_fill
        else:
            if row_data['D']:
                if row_data['D_has_L']:
                    ws[f'E{row_num}'].fill = ai_generated_fill
                else:
                    ws[f'E{row_num}'].fill = pending_fill
            if row_data['F']:
                if row_data['F_has_L']:
                    ws[f'G{row_num}'].fill = ai_generated_fill
                else:
                    ws[f'G{row_num}'].fill = pending_fill
            if row_data['H']:
                if row_data['H_has_L']:
                    ws[f'I{row_num}'].fill = ai_generated_fill
                else:
                    ws[f'I{row_num}'].fill = pending_fill
    
    file_name = os.path.basename(excel_path)
    prefix = file_name.split('_preview')[0] if '_preview' in file_name else file_name.replace('.xlsx', '')
    timestamp = time.strftime('%Y%m%d_%H%M%S')
    output_path = os.path.join(out_dir, f'{prefix}_complete_{timestamp}.xlsx')
    wb.save(output_path)
    
    print(f'处理完成！输出文件: {output_path}')
    print(f'规范总需求项数: {len(all_spec_ids)}')
    print(f'原始数据行数(含重复): {len(all_original_rows)}')
    print(f'输出总行数: {len(new_rows_data)}')
    
    new_count = sum(1 for r in new_rows_data if r['is_new'])
    print(f'新增行数: {new_count}')
    
    print(f'原始独有的ID: {original_only_ids}')

def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(script_dir)
    
    data_dir = os.path.join(project_root, 'complete_data')
    json_path = os.path.join(project_root, 'data', 'config', '芯片需求结构化定义规范V4.0.json')
    out_dir = os.path.join(project_root, 'conplete_out')
    
    os.makedirs(out_dir, exist_ok=True)
    
    with open(json_path, 'r', encoding='utf-8') as f:
        spec_data = json.load(f)
    
    spec_map = {item['需求项序号']: item for item in spec_data}
    
    excel_files = [f for f in os.listdir(data_dir) if f.endswith('.xlsx')]
    
    if not excel_files:
        print('未找到任何Excel文件！')
        return
    
    print(f'找到 {len(excel_files)} 个Excel文件: {excel_files}')
    
    for excel_file in excel_files:
        excel_path = os.path.join(data_dir, excel_file)
        print(f'\n正在处理: {excel_file}')
        process_excel_file(excel_path, spec_map, out_dir)

if __name__ == '__main__':
    main()